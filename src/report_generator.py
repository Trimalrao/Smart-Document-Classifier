"""
Report Generator  (fixed)
=========================
Groups classified + extracted documents and writes a styled Excel report.

Fixes applied:
  BANK  - Grouping uses normalized account number as primary key.
          If account number is missing, falls back to normalized holder name.
  PAY   - Groups strictly by Employee ID (not name).
          If Employee ID missing, falls back to cleaned Employee Name.
  TAX   - Groups strictly by Employee PAN (never employer PAN).
          Falls back to employee name only when PAN is truly absent.

Sheets:
  Summary | Bank Statements | Payslips | Tax Documents
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

from .entity_extractor import extract_entities, month_sort_key

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False


# ─────────────────────────────────────────────────────────────────────────────
# Normalisation helpers
# ─────────────────────────────────────────────────────────────────────────────

def _norm_account(raw: str) -> str:
    """Strip spaces, dashes, X/x masks → pure digit string or masked form."""
    if not raw:
        return ""
    cleaned = re.sub(r"[\s\-]", "", raw).upper()
    return cleaned if cleaned else ""


def _norm_name(raw: str) -> str:
    """Lowercase, collapse spaces — used as fallback key when ID is missing."""
    return " ".join(raw.lower().split()) if raw else ""


# ─────────────────────────────────────────────────────────────────────────────
# Grouping logic
# ─────────────────────────────────────────────────────────────────────────────

def _group_bank_statements(records: list[dict]) -> list[dict]:
    """
    Group by normalised account number.
    Falls back to normalised holder name when account number is absent.
    """
    groups: dict[str, list[dict]] = defaultdict(list)

    for r in records:
        acct = _norm_account(r.get("account_number", ""))
        if acct:
            key = acct
        else:
            # Fallback: use normalised holder name so at least the
            # same person's statements land together
            name = _norm_name(r.get("holder_name", ""))
            key  = name if name else "UNKNOWN"
        groups[key].append(r)

    rows = []
    for key, items in groups.items():
        items_sorted = sorted(
            items,
            key=lambda x: month_sort_key(x.get("statement_month", "")),
            reverse=True,
        )
        latest = items_sorted[0]

        # Pick best (longest) non-empty holder name from all items in group
        names  = [i.get("holder_name", "") for i in items if i.get("holder_name", "").strip()]
        holder = max(names, key=len) if names else ""

        banks  = [i.get("bank_name", "") for i in items if i.get("bank_name", "").strip()]
        bank   = banks[0] if banks else ""

        # Display account number: use the raw (non-normalised) value from latest doc
        display_acct = latest.get("account_number", "") or key

        rows.append({
            "Account Number":         display_acct,
            "Account Holder":         holder,
            "Bank Name":              bank,
            "Latest Statement Month": latest.get("statement_month", ""),
            "Latest PDF File Name":   Path(latest.get("filename", "")).name,
            "Total Statements":       len(items),
            "All Months":             ", ".join(
                i.get("statement_month", "—") for i in items_sorted
            ),
        })

    return sorted(rows, key=lambda x: x["Account Number"])


def _group_payslips(records: list[dict]) -> list[dict]:
    """
    Group strictly by Employee ID.
    Falls back to normalised Employee Name when ID is absent.
    """
    groups: dict[str, list[dict]] = defaultdict(list)

    for r in records:
        emp_id = r.get("employee_id", "").strip()
        if emp_id:
            key = emp_id.upper()
        else:
            name = _norm_name(r.get("employee_name", ""))
            key  = name if name else "UNKNOWN"
        groups[key].append(r)

    rows = []
    for key, items in groups.items():
        items_sorted = sorted(
            items,
            key=lambda x: month_sort_key(x.get("payslip_month", "")),
            reverse=True,
        )
        latest = items_sorted[0]

        names    = [i.get("employee_name", "") for i in items if i.get("employee_name", "").strip()]
        emp_name = max(names, key=len) if names else ""

        companies = [i.get("company", "") for i in items if i.get("company", "").strip()]
        company   = companies[0] if companies else ""

        emp_id_display = latest.get("employee_id", "") or key

        rows.append({
            "Employee ID":          emp_id_display,
            "Employee Name":        emp_name,
            "Company":              company,
            "Latest Payslip Month": latest.get("payslip_month", ""),
            "Latest PDF File Name": Path(latest.get("filename", "")).name,
            "Total Payslips":       len(items),
            "All Months":           ", ".join(
                i.get("payslip_month", "—") for i in items_sorted
            ),
        })

    return sorted(rows, key=lambda x: x["Employee ID"])


def _group_tax_documents(records: list[dict]) -> list[dict]:
    """
    Group strictly by Employee PAN.
    Falls back to normalised employee name when PAN is absent.
    Employer PAN is NEVER used as a grouping key.
    """
    groups: dict[str, list[dict]] = defaultdict(list)

    for r in records:
        pan = r.get("pan", "").strip().upper()
        if pan:
            key = pan
        else:
            name = _norm_name(r.get("employee_name", ""))
            key  = name if name else "UNKNOWN"
        groups[key].append(r)

    rows = []
    for key, items in groups.items():
        # Sort by assessment year descending
        def ay_sort(x):
            ay = x.get("assessment_year", "")
            m  = re.search(r"(\d{4})", ay)
            return int(m.group(1)) if m else 0

        items_sorted = sorted(items, key=ay_sort, reverse=True)
        latest       = items_sorted[0]

        names    = [i.get("employee_name", "") for i in items if i.get("employee_name", "").strip()]
        emp_name = max(names, key=len) if names else ""

        employers = [i.get("employer", "") for i in items if i.get("employer", "").strip()]
        employer  = employers[0] if employers else ""

        pan_display = latest.get("pan", "") or key

        rows.append({
            "PAN Number":             pan_display,
            "Employee Name":          emp_name,
            "Employer":               employer,
            "Latest Assessment Year": latest.get("assessment_year", ""),
            "File Name":              Path(latest.get("filename", "")).name,
            "Total Documents":        len(items),
            "All Assessment Years":   ", ".join(
                i.get("assessment_year", "—") for i in items_sorted
            ),
        })

    return sorted(rows, key=lambda x: x["PAN Number"])


# ─────────────────────────────────────────────────────────────────────────────
# Excel styling
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1A1A2E")
_SUBHEAD_FILL = PatternFill("solid", fgColor="E8E8F0")
_ALT_FILL     = PatternFill("solid", fgColor="F4F4FA")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_BODY_FONT    = Font(size=9)
_BOLD_FONT    = Font(bold=True, size=9)
_THIN         = Side(style="thin", color="CCCCCC")
_THIN_BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _style_sheet(ws, df, title: str, subtitle: str):
    ws.insert_rows(1, 2)
    ncols = len(df.columns)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = title
    c.font  = Font(bold=True, size=13, color="FFFFFF")
    c.fill  = _HEADER_FILL
    c.alignment = _CENTER
    ws.row_dimensions[1].height = 22

    # Subtitle row
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]
    c.value = subtitle
    c.font  = Font(italic=True, size=9, color="444466")
    c.fill  = _SUBHEAD_FILL
    c.alignment = _CENTER
    ws.row_dimensions[2].height = 16

    # Header row (row 3)
    for cell in ws[3]:
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _THIN_BORDER
    ws.row_dimensions[3].height = 18

    # Data rows
    for i, row in enumerate(ws.iter_rows(min_row=4), start=1):
        fill = _ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for cell in row:
            cell.font      = _BODY_FONT
            cell.fill      = fill
            cell.alignment = _LEFT
            cell.border    = _THIN_BORDER

    # Auto column widths
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 50)

    ws.freeze_panes = "A4"


def _add_summary_sheet(wb, counts: dict, total: int):
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    data = [
        ("DocSense — Document Summary Report", None),
        ("", None),
        ("Category", "Count"),
        ("Total Documents Processed", total),
        ("", None),
        ("Bank Statements",                 counts.get("Bank Statement", 0)),
        ("Payslips",                        counts.get("Payslip", 0)),
        ("Tax Documents",                   counts.get("Tax Document", 0)),
        ("Others (excluded from report)",   counts.get("Others", 0)),
    ]
    for r_idx, (label, value) in enumerate(data, 1):
        ws.cell(row=r_idx, column=1, value=label)
        if value is not None:
            ws.cell(row=r_idx, column=2, value=value)
        if r_idx == 1:
            ws.cell(row=r_idx, column=1).font = Font(bold=True, size=14)
            ws.row_dimensions[r_idx].height = 24
        elif r_idx == 3:
            for col in (1, 2):
                c = ws.cell(row=r_idx, column=col)
                c.font = _HEADER_FONT
                c.fill = _HEADER_FILL
                c.alignment = _CENTER
        elif r_idx in (4, 6, 7, 8, 9):
            ws.cell(row=r_idx, column=1).font = _BODY_FONT
            c2 = ws.cell(row=r_idx, column=2)
            c2.font      = _BOLD_FONT
            c2.alignment = _CENTER
            if r_idx in (6, 7, 8):
                ws.cell(row=r_idx, column=1).fill = _ALT_FILL
                c2.fill = _ALT_FILL
    ws.freeze_panes = "A1"


# ─────────────────────────────────────────────────────────────────────────────
# Main entry point
# ─────────────────────────────────────────────────────────────────────────────

def generate_report(
    pipeline_results: list,
    output_path: str = "document_summary.xlsx",
) -> str:
    """
    Extract entities from classified results, group them, write Excel.

    Args:
        pipeline_results : list[PipelineResult] from DocumentClassificationPipeline
        output_path      : output .xlsx file path

    Returns:
        Resolved absolute path of the written file.
    """
    if not _EXCEL_AVAILABLE:
        raise ImportError("Run: pip install pandas openpyxl")

    bank_records: list[dict] = []
    pay_records:  list[dict] = []
    tax_records:  list[dict] = []
    counts:       dict[str, int] = {}

    for r in pipeline_results:
        label = r.classification.label
        counts[label] = counts.get(label, 0) + 1

        if label in ("Bank Statement", "Payslip", "Tax Document"):
            entities = extract_entities(label, r.cleaned_text, r.file_path)
            if label == "Bank Statement":
                bank_records.append(entities)
            elif label == "Payslip":
                pay_records.append(entities)
            elif label == "Tax Document":
                tax_records.append(entities)

    bank_rows = _group_bank_statements(bank_records)
    pay_rows  = _group_payslips(pay_records)
    tax_rows  = _group_tax_documents(tax_records)

    _BANK_COLS = ["Account Number", "Account Holder", "Bank Name",
                  "Latest Statement Month", "Latest PDF File Name",
                  "Total Statements", "All Months"]
    _PAY_COLS  = ["Employee ID", "Employee Name", "Company",
                  "Latest Payslip Month", "Latest PDF File Name",
                  "Total Payslips", "All Months"]
    _TAX_COLS  = ["PAN Number", "Employee Name", "Employer",
                  "Latest Assessment Year", "File Name",
                  "Total Documents", "All Assessment Years"]

    bank_df = pd.DataFrame(bank_rows) if bank_rows else pd.DataFrame(columns=_BANK_COLS)
    pay_df  = pd.DataFrame(pay_rows)  if pay_rows  else pd.DataFrame(columns=_PAY_COLS)
    tax_df  = pd.DataFrame(tax_rows)  if tax_rows  else pd.DataFrame(columns=_TAX_COLS)

    out = Path(output_path)
    total = sum(counts.values())

    with pd.ExcelWriter(str(out), engine="openpyxl") as writer:
        bank_df.to_excel(writer, sheet_name="Bank Statements", index=False)
        pay_df.to_excel(writer,  sheet_name="Payslips",        index=False)
        tax_df.to_excel(writer,  sheet_name="Tax Documents",   index=False)

    wb = load_workbook(str(out))

    _style_sheet(
        wb["Bank Statements"], bank_df,
        "Bank Statements Summary",
        f"{len(bank_rows)} unique accounts from {counts.get('Bank Statement', 0)} statement(s)",
    )
    _style_sheet(
        wb["Payslips"], pay_df,
        "Payslips Summary",
        f"{len(pay_rows)} unique employees from {counts.get('Payslip', 0)} payslip(s)",
    )
    _style_sheet(
        wb["Tax Documents"], tax_df,
        "Tax Documents Summary",
        f"{len(tax_rows)} unique PAN numbers from {counts.get('Tax Document', 0)} document(s)",
    )
    _add_summary_sheet(wb, counts, total)

    wb.save(str(out))
    return str(out.resolve())